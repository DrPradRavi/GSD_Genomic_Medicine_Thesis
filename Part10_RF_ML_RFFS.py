import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import RepeatedStratifiedKFold
from sklearn.metrics import roc_auc_score, accuracy_score, f1_score, precision_score, recall_score
from sklearn.preprocessing import OrdinalEncoder
import os
from joblib import Parallel, delayed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import multiprocessing

# Configuration variables
N_PERMUTATIONS = 1000
available_cores = multiprocessing.cpu_count()
n_jobs = max(1, available_cores - 6)
print(f"Using {n_jobs} cores for parallel processing.")

# File paths
input_csv_selected_features = "Part9_RF_feature_selection_results/Part9_selected_features.csv"
output_dir = "Part10_RF_output_RF"
os.makedirs(output_dir, exist_ok=True)

# Load the dataset with selected features
print("Starting task: Load the dataset with selected features")
df = pd.read_csv(input_csv_selected_features, index_col=0).reset_index(drop=True)
print("Finished task: Load the dataset with selected features")

# Separate features and target
print("Starting task: Separate features and target")
X = df.iloc[:, :-1]
y = df.iloc[:, -1].astype(int)
print("Finished task: Separate features and target")

# Encode categorical features
print("Starting task: Encode categorical features")
def process_column(column):
    if X[column].dtype == 'object':
        return pd.factorize(X[column])[0]
    return X[column].values

X = pd.DataFrame(Parallel(n_jobs=n_jobs)(delayed(process_column)(column) for column in X.columns), index=X.columns).T
print("Finished task: Encode categorical features")

# Initialize results storage
results = []

# Define cross-validation
cv = RepeatedStratifiedKFold(n_splits=5, n_repeats=3)

# Parameter grid
mtry_values = [10, 50, 100, 300]  
ntree_values = [10, 50, 100, 300] 
max_depth_values = [2, 5, 10]

# Function to perform cross-validation for a given set of parameters
def evaluate_model(mtry, ntree, max_depth):
    fold_accuracies, fold_aucs, fold_f1s, fold_precisions, fold_recalls = [], [], [], [], []

    for train_index, test_index in cv.split(X, y):
        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        model = RandomForestClassifier(n_estimators=ntree, max_features=mtry, max_depth=max_depth)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_pred_prob = model.predict_proba(X_test)[:, 1]

        fold_accuracies.append(accuracy_score(y_test, y_pred))
        fold_aucs.append(roc_auc_score(y_test, y_pred_prob))
        fold_f1s.append(f1_score(y_test, y_pred))
        fold_precisions.append(precision_score(y_test, y_pred))
        fold_recalls.append(recall_score(y_test, y_pred))

    avg_accuracy = np.mean(fold_accuracies)
    avg_auc = np.mean(fold_aucs)
    avg_f1 = np.mean(fold_f1s)
    avg_precision = np.mean(fold_precisions)
    avg_recall = np.mean(fold_recalls)

    return (mtry, ntree, max_depth, fold_accuracies, fold_aucs, fold_f1s, fold_precisions, fold_recalls, 
            avg_accuracy, avg_auc, avg_f1, avg_precision, avg_recall)

# Perform grid search in parallel
print("Starting task: Perform grid search")
total_configs = len(mtry_values) * len(ntree_values) * len(max_depth_values)
results = []

for i, (mtry, ntree, max_depth) in enumerate([(m, n, d) for m in mtry_values for n in ntree_values for d in max_depth_values], 1):
    result = evaluate_model(mtry, ntree, max_depth)
    results.append(result)
    print(f"{i}/{total_configs} parameters have been completed")

print("Finished task: Perform grid search")

# Convert results to DataFrame
print("Starting task: Convert results to DataFrame")
results_df = pd.DataFrame(results, columns=['mtry', 'ntree', 'max_depth', 'Fold Accuracies', 'Fold AUCs', 'Fold F1s', 'Fold Precisions', 'Fold Recalls',
                                            'Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall'])
detailed_results_df = results_df.copy()
print("Finished task: Convert results to DataFrame")

# Save detailed results to CSV
print("Starting task: Save detailed results to CSV")
for col in ['Fold Accuracies', 'Fold AUCs', 'Fold F1s', 'Fold Precisions', 'Fold Recalls']:
    detailed_results_df[col] = detailed_results_df[col].apply(lambda x: [float(i) for i in x])
detailed_results_df.to_csv(os.path.join(output_dir, 'Part10_RF_detailed_results.csv'), index=False)
print("Finished task: Save detailed results to CSV")

# Save summary results as an Excel file
print("Starting task: Save summary results as an Excel file")
summary_df = results_df[['mtry', 'ntree', 'max_depth', 'Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall']].copy()
summary_df['mtry'] = summary_df['mtry'].astype(int)
summary_df['ntree'] = summary_df['ntree'].astype(int)
summary_df['max_depth'] = summary_df['max_depth'].astype(int)
for col in ['Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall']:
    summary_df[col] = summary_df[col].astype(float).apply(lambda x: f"{x:.4f}")

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Summary Results"

# Write the DataFrame to the worksheet
for r_idx, row in enumerate(summary_df.itertuples(), 1):
    for c_idx, value in enumerate(row[1:], 1):
        ws.cell(row=r_idx + 1, column=c_idx, value=value)

# Write the header
for c_idx, col_name in enumerate(summary_df.columns, 1):
    ws.cell(row=1, column=c_idx, value=col_name)

# Highlight the row with the highest Avg AUC
max_auc_idx = summary_df['Avg AUC'].astype(float).idxmax() + 2  # +2 to account for header and 1-based index
fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
for col in range(1, len(summary_df.columns) + 1):
    ws.cell(row=max_auc_idx, column=col).fill = fill

# Save the workbook
wb.save(os.path.join(output_dir, 'Part10_RF_summary.xlsx'))
print("Finished task: Save summary results as an Excel file")

# Define color palette
colors = [
    (239/255, 45/255, 81/255),  # Red
    (255/255, 221/255, 0/255),  # Yellow
    (33/255, 150/255, 243/255),  # Royal Blue
    (76/255, 175/255, 80/255),  # Green
    (255/255, 152/255, 0/255),  # Orange
    (156/255, 39/255, 176/255)  # Purple
]

# Define markers and linestyles
markers = ['o', 's', 'D', 'v', '<', '>', 'p', '*', 'h']  # More interesting symbols
linestyles = ['-', '--', '-.', ':']  # Different line styles

# Plot average accuracy and AUC as ntree and mtry change
print("Starting task: Plot average metrics as ntree and mtry change")
metrics = [('Avg Accuracy', 'Average Accuracy'), ('Avg AUC', 'Average AUC'), 
           ('Avg F1', 'Average F1'), ('Avg Precision', 'Average Precision'), ('Avg Recall', 'Average Recall')]
fig, axs = plt.subplots(len(max_depth_values), len(metrics), figsize=(24 * len(metrics) // 2, 8 * len(max_depth_values)))

for row, max_depth in enumerate(max_depth_values):
    for col, (metric, ylabel) in enumerate(metrics):
        ax = axs[row, col]
        for i, mtry in enumerate(mtry_values):
            subset = summary_df[(summary_df['max_depth'] == max_depth) & (summary_df['mtry'] == mtry)]
            ax.plot(subset['ntree'], subset[metric].astype(float), label=f'mtry={mtry}', color=colors[i % len(colors)], marker=markers[i % len(markers)], markersize=8, markerfacecolor='white', markeredgewidth=2, linestyle=linestyles[i % len(linestyles)])
        ax.set_xlabel('Number of Trees')
        ax.set_ylabel(ylabel)
        ax.set_title(f'RF {ylabel} vs Number of Trees (max_depth={max_depth})')
        ax.legend(loc='best')
        ax.grid(True, which='both', linestyle='--', linewidth=0.5)
        ax.axhline(y=0, color='black', linewidth=0.5, alpha=0.5)

fig.tight_layout()
plt.savefig(os.path.join(output_dir, 'Part10_RF_combined_plots.png'))
print("Finished task: Plot average metrics as ntree and mtry change")

# Identify best parameters
print("Starting task: Identify best parameters")
best_params = summary_df.loc[summary_df['Avg AUC'].astype(float).idxmax()]
best_mtry = int(best_params['mtry'])
best_ntree = int(best_params['ntree'])
best_max_depth = int(best_params['max_depth'])
print("Finished task: Identify best parameters")

# Train final model with best parameters and compute Gini importance
print("Starting task: Train final model and compute Gini importance")
start_time = time.time()

n_repeats = 100  # Reduced from 1000 to 100
gini_importances = []
cv = RepeatedStratifiedKFold(n_splits=5, n_repeats=3)

for i in range(n_repeats):
    fold_importances = []
    for train_index, test_index in cv.split(X, y):
        X_train, X_test = X.iloc[train_index], X.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]
        
        final_model = RandomForestClassifier(n_estimators=best_ntree, max_features=best_mtry, max_depth=best_max_depth)
        final_model.fit(X_train, y_train)
        fold_importances.append(final_model.feature_importances_)
    
    gini_importances.append(np.mean(fold_importances, axis=0))
    
    # Print progress every 10 iterations
    if (i + 1) % 10 == 0:
        current_time = time.time()
        elapsed_time = current_time - start_time
        print(f"Completed {i + 1}/{n_repeats} iterations in {elapsed_time:.2f} seconds")

gini_importances = np.array(gini_importances)
avg_gini_importance = np.mean(gini_importances, axis=0)
std_gini_importance = np.std(gini_importances, axis=0)
p_values = np.mean(gini_importances >= avg_gini_importance[np.newaxis, :], axis=0)

end_time = time.time()
training_time = end_time - start_time
print(f"Finished task: Train final model and compute Gini importance in {training_time:.2f} seconds")

# Create and save feature importance DataFrame
gini_importance_df = pd.DataFrame({
    'feature': X.columns,
    'importance': avg_gini_importance,
    'std': std_gini_importance,
    'p_value': p_values
}).sort_values('importance', ascending=False)
gini_importance_df['rank'] = range(1, len(gini_importance_df) + 1)

# Save all features' importance to CSV
print("Starting task: Save all features' importance to CSV")
gini_importance_df.to_csv(os.path.join(output_dir, 'Part10_RF_all_features_importance.csv'), index=False)
print("Finished task: Save all features' importance to CSV")

# Select top 30 features
print("Starting task: Select top 30 features")
top_features = gini_importance_df.head(30).copy()
top_features['rank'] = range(1, 31)
print("Finished task: Select top 30 features")

# Save top features to CSV
print("Starting task: Save top features to CSV")
top_features.to_csv(os.path.join(output_dir, 'Part10_RF_top_features_importance.csv'), index=False)
print("Finished task: Save top features to CSV")

# Save a version of the input CSV with only the top 30 features
print("Starting task: Save input CSV with top 30 features")
top_feature_names = top_features['feature'].tolist()

# Load the original selected features CSV to get the correct index
original_df = pd.read_csv("Part9_RF_feature_selection_results/Part9_selected_features.csv", index_col=0)

# Select top 30 features and the target column
df_top_features = original_df[top_feature_names + [original_df.columns[-1]]]

# Ensure the index name is 'index_name'
df_top_features.index.name = 'index_name'

# Save the CSV file
df_top_features.to_csv(os.path.join(output_dir, 'Part10_RF_input_top_30_features.csv'))
print("Finished task: Save input CSV with top 30 features")

# Create histogram of feature importance scores
print("Starting task: Create histogram of feature importance scores")
plt.figure(figsize=(12, 6))

# Define colors
red = (239/255, 45/255, 81/255)
purple = (156/255, 39/255, 176/255)
orange = (255/255, 152/255, 0/255)
yellow = (255/255, 221/255, 0/255)

plt.hist(gini_importance_df['importance'], bins=50, edgecolor='black', color=red, alpha=0.7)
plt.xscale('symlog', linthresh=1e-5)  # Symmetric log scale
plt.xlabel('Feature Importance (Gini)')
plt.ylabel('Frequency')
plt.title('Distribution of Feature Importance Scores')

mean_importance = gini_importance_df['importance'].mean()
median_importance = gini_importance_df['importance'].median()
top_30_threshold = gini_importance_df['importance'].nlargest(30).min()

plt.axvline(x=mean_importance, color=purple, linestyle='dashed', linewidth=2, label='Mean Importance')
plt.axvline(x=median_importance, color=orange, linestyle='dashed', linewidth=2, label='Median Importance')
plt.axvline(x=top_30_threshold, color=yellow, linestyle='dashed', linewidth=2, label='Top 30 Threshold')

plt.legend()
plt.grid(True, which="both", ls="-", alpha=0.2)
plt.tight_layout()
plt.savefig(os.path.join(output_dir, 'Part10_RF_feature_importance_histogram.png'))
plt.close()
print("Finished task: Create histogram of feature importance scores")

print('Random Forest with Gini importance calculation completed')