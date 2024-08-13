import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from xgboost import XGBClassifier
from sklearn.model_selection import RepeatedStratifiedKFold
from sklearn.metrics import roc_auc_score, accuracy_score, f1_score, precision_score, recall_score
import os
from joblib import Parallel, delayed
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import time
import multiprocessing

# Configuration variables
N_PERMUTATIONS = 1000
available_cores = multiprocessing.cpu_count()
n_jobs = max(1, available_cores - 6)
print(f"Using {n_jobs} cores for parallel processing.")

# File paths
input_csv_selected_features = "Part9_RF_feature_selection_results/Part9_selected_features.csv"
output_dir = "Part11_XGB_output_RF"
os.makedirs(output_dir, exist_ok=True)

# Load the dataset with selected features
print("Starting task: Load the dataset with selected features")
df = pd.read_csv(input_csv_selected_features, index_col=0).reset_index(drop=True)
print("Finished task: Load the dataset with selected features")

# Separate features and target
print("Starting task: Separate features and target")
X = df.iloc[:, :-1]
y = df.iloc[:, -1].astype(int)
print("Finished task: Separate features and target")

# Encode categorical features using pandas get_dummies
print("Starting task: Encode categorical features")
X_encoded = pd.get_dummies(X, drop_first=True)
print("Finished task: Encode categorical features")

# Get feature names after encoding
feature_names = X_encoded.columns.tolist()

# Initialize results storage
results = []

# Define cross-validation
cv = RepeatedStratifiedKFold(n_splits=5, n_repeats=3)

# Parameter grid

learning_rate_values = [0.01, 0.05, 0.1, 0.3]
n_estimators_values = [10, 50, 100, 300] 
max_depth_values = [2, 5, 10] 
# Function to perform cross-validation for a given set of parameters
def evaluate_model(learning_rate, max_depth, n_estimators):
    fold_accuracies, fold_aucs, fold_f1s, fold_precisions, fold_recalls = [], [], [], [], []

    for train_index, test_index in cv.split(X_encoded, y):
        X_train, X_test = X_encoded.iloc[train_index], X_encoded.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        model = XGBClassifier(learning_rate=learning_rate, max_depth=max_depth, n_estimators=n_estimators, eval_metric='logloss')
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        y_pred_prob = model.predict_proba(X_test)[:, 1]

        fold_accuracies.append(accuracy_score(y_test, y_pred))
        fold_aucs.append(roc_auc_score(y_test, y_pred_prob))
        fold_f1s.append(f1_score(y_test, y_pred))
        fold_precisions.append(precision_score(y_test, y_pred))
        fold_recalls.append(recall_score(y_test, y_pred))

    avg_accuracy = np.mean(fold_accuracies)
    avg_auc = np.mean(fold_aucs)
    avg_f1 = np.mean(fold_f1s)
    avg_precision = np.mean(fold_precisions)
    avg_recall = np.mean(fold_recalls)

    return (learning_rate, max_depth, n_estimators, fold_accuracies, fold_aucs, fold_f1s, fold_precisions, fold_recalls, 
            avg_accuracy, avg_auc, avg_f1, avg_precision, avg_recall)

# Perform grid search in parallel
print("Starting task: Perform grid search")
total_configs = len(learning_rate_values) * len(max_depth_values) * len(n_estimators_values)
results = []

for i, (learning_rate, max_depth, n_estimators) in enumerate([(lr, d, n) for lr in learning_rate_values for d in max_depth_values for n in n_estimators_values], 1):
    result = evaluate_model(learning_rate, max_depth, n_estimators)
    results.append(result)
    print(f"{i}/{total_configs} parameters have been completed")

print("Finished task: Perform grid search")

# Convert results to DataFrame
print("Starting task: Convert results to DataFrame")
results_df = pd.DataFrame(results, columns=['learning_rate', 'max_depth', 'n_estimators', 'Fold Accuracies', 'Fold AUCs', 'Fold F1s', 'Fold Precisions', 'Fold Recalls',
                                            'Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall'])
detailed_results_df = results_df.copy()
print("Finished task: Convert results to DataFrame")

# Save detailed results to CSV
print("Starting task: Save detailed results to CSV")
for col in ['Fold Accuracies', 'Fold AUCs', 'Fold F1s', 'Fold Precisions', 'Fold Recalls']:
    detailed_results_df[col] = detailed_results_df[col].apply(lambda x: [float(i) for i in x])
detailed_results_df.to_csv(os.path.join(output_dir, 'Part11_XGB_detailed_results.csv'), index=False)
print("Finished task: Save detailed results to CSV")

# Save summary results as an Excel file
print("Starting task: Save summary results as an Excel file")
summary_df = results_df[['learning_rate', 'max_depth', 'n_estimators', 'Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall']].copy()
summary_df['max_depth'] = summary_df['max_depth'].astype(int)
summary_df['n_estimators'] = summary_df['n_estimators'].astype(int)
for col in ['Avg Accuracy', 'Avg AUC', 'Avg F1', 'Avg Precision', 'Avg Recall']:
    summary_df[col] = summary_df[col].astype(float).apply(lambda x: f"{x:.4f}")

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Summary Results"

# Write the DataFrame to the worksheet
for r_idx, row in enumerate(summary_df.itertuples(), 1):
    for c_idx, value in enumerate(row[1:], 1):
        ws.cell(row=r_idx + 1, column=c_idx, value=value)

# Write the header
for c_idx, col_name in enumerate(summary_df.columns, 1):
    ws.cell(row=1, column=c_idx, value=col_name)

# Highlight the row with the highest Avg AUC
max_auc_idx = summary_df['Avg AUC'].astype(float).idxmax() + 2  # +2 to account for header and 1-based index
fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
for col in range(1, len(summary_df.columns) + 1):
    ws.cell(row=max_auc_idx, column=col).fill = fill

# Save the workbook
wb.save(os.path.join(output_dir, 'Part11_XGB_summary.xlsx'))
print("Finished task: Save summary results as an Excel file")

# Define color palette
colors = [
    (239/255, 45/255, 81/255),  # Red
    (255/255, 221/255, 0/255),  # Yellow
    (33/255, 150/255, 243/255),  # Royal Blue
    (76/255, 175/255, 80/255),  # Green
    (255/255, 152/255, 0/255),  # Orange
    (156/255, 39/255, 176/255)  # Purple
]

# Define markers and linestyles
markers = ['o', 's', 'D', 'v', '<', '>', 'p', '*', 'h']  # More interesting symbols
linestyles = ['-', '--', '-.', ':']  # Different line styles

# Plot average accuracy and AUC as n_estimators and learning_rate change
print("Starting task: Plot average metrics as n_estimators and learning_rate change")
metrics = [('Avg Accuracy', 'Average Accuracy'), ('Avg AUC', 'Average AUC'), 
           ('Avg F1', 'Average F1'), ('Avg Precision', 'Average Precision'), ('Avg Recall', 'Average Recall')]
fig, axs = plt.subplots(len(max_depth_values), len(metrics), figsize=(24 * len(metrics) // 2, 8 * len(max_depth_values)))

for row, max_depth in enumerate(max_depth_values):
    for col, (metric, ylabel) in enumerate(metrics):
        ax = axs[row, col]
        for i, learning_rate in enumerate(learning_rate_values):
            subset = summary_df[(summary_df['max_depth'] == max_depth) & (summary_df['learning_rate'] == learning_rate)]
            ax.plot(subset['n_estimators'], subset[metric].astype(float), label=f'learning_rate={learning_rate}', color=colors[i % len(colors)], marker=markers[i % len(markers)], markersize=8, markerfacecolor='white', markeredgewidth=2, linestyle=linestyles[i % len(linestyles)])
        ax.set_xlabel('Number of Estimators')
        ax.set_ylabel(ylabel)
        ax.set_title(f'XGBoost {ylabel} vs Number of Estimators (max_depth={max_depth})')
        ax.legend(loc='best')
        ax.grid(True, which='both', linestyle='--', linewidth=0.5)
        ax.axhline(y=0, color='black', linewidth=0.5, alpha=0.5)

fig.tight_layout()
plt.savefig(os.path.join(output_dir, 'Part11_XGB_combined_plots.png'))
print("Finished task: Plot average metrics as n_estimators and learning_rate change")

# Identify best parameters
print("Starting task: Identify best parameters")
best_params = summary_df.loc[summary_df['Avg AUC'].astype(float).idxmax()]
best_learning_rate = float(best_params['learning_rate'])
best_max_depth = int(best_params['max_depth'])
best_n_estimators = int(best_params['n_estimators'])
print("Finished task: Identify best parameters")

# Train final model with best parameters and compute feature importance
print("Starting task: Train final model and compute feature importance")
start_time = time.time()

n_repeats = 100  # Reduced from 100 to 10
feature_importances = []
cv_scores = {'accuracy': [], 'auc': [], 'f1': [], 'precision': [], 'recall': []}

for i in range(n_repeats):
    fold_importances = []
    for train_index, test_index in cv.split(X_encoded, y):
        X_train, X_test = X_encoded.iloc[train_index], X_encoded.iloc[test_index]
        y_train, y_test = y.iloc[train_index], y.iloc[test_index]

        final_model = XGBClassifier(learning_rate=best_learning_rate, max_depth=best_max_depth, n_estimators=best_n_estimators, eval_metric='logloss')
        final_model.fit(X_train, y_train)
        
        # Compute feature importance for this fold
        fold_importances.append(final_model.feature_importances_)
        
        # Compute performance metrics for this fold
        y_pred = final_model.predict(X_test)
        y_pred_prob = final_model.predict_proba(X_test)[:, 1]
        
        cv_scores['accuracy'].append(accuracy_score(y_test, y_pred))
        cv_scores['auc'].append(roc_auc_score(y_test, y_pred_prob))
        cv_scores['f1'].append(f1_score(y_test, y_pred))
        cv_scores['precision'].append(precision_score(y_test, y_pred))
        cv_scores['recall'].append(recall_score(y_test, y_pred))
    
    # Average feature importances across folds for this repeat
    feature_importances.append(np.mean(fold_importances, axis=0))
    
    # Print progress every 10 iterations
    if (i + 1) % 10 == 0:
        current_time = time.time()
        elapsed_time = current_time - start_time
        print(f"Completed {i + 1}/{n_repeats} iterations in {elapsed_time:.2f} seconds")

feature_importances = np.array(feature_importances)
avg_feature_importance = np.mean(feature_importances, axis=0)
std_feature_importance = np.std(feature_importances, axis=0)
p_values = np.mean(feature_importances >= avg_feature_importance[np.newaxis, :], axis=0)

end_time = time.time()
training_time = end_time - start_time
print(f"Finished task: Train final model and compute feature importance in {training_time:.2f} seconds")

# Print average CV scores
print("Average CV scores:")
for metric, scores in cv_scores.items():
    print(f"{metric.capitalize()}: {np.mean(scores):.4f} (+/- {np.std(scores) * 2:.4f})")

# Create and save feature importance DataFrame
feature_importance_df = pd.DataFrame({
    'feature': feature_names,
    'importance': avg_feature_importance,
    'std': std_feature_importance,
    'p_value': p_values
}).sort_values('importance', ascending=False)
feature_importance_df['rank'] = range(1, len(feature_importance_df) + 1)

# Save all features' importance to CSV
print("Starting task: Save all features' importance to CSV")
feature_importance_df.to_csv(os.path.join(output_dir, 'Part11_XGB_all_features_importance.csv'), index=False)
print("Finished task: Save all features' importance to CSV")

# Select top 30 features
print("Starting task: Select top 30 features")
top_features = feature_importance_df.head(30).copy()
top_features['rank'] = range(1, 31)
print("Finished task: Select top 30 features")

# Save top features to CSV
print("Starting task: Save top features to CSV")
top_features.to_csv(os.path.join(output_dir, 'Part11_XGB_top_features_importance.csv'), index=False)
print("Finished task: Save top features to CSV")

# Save a version of the input CSV with only the top 30 features
print("Starting task: Save input CSV with top 30 features")
top_feature_names = top_features['feature'].tolist()

# Load the original selected features CSV to get the correct index
original_df = pd.read_csv("Part9_RF_feature_selection_results/Part9_selected_features.csv", index_col=0)

# Select top 30 features and the target column
df_top_features = original_df[top_feature_names + [original_df.columns[-1]]]

# Ensure the index name is 'index_name'
df_top_features.index.name = 'index_name'

# Save the CSV file
df_top_features.to_csv(os.path.join(output_dir, 'Part11_XGB_input_top_30_features.csv'))
print("Finished task: Save input CSV with top 30 features")

# Create histogram of feature importance scores
print("Starting task: Create histogram of feature importance scores")
plt.figure(figsize=(12, 6))

# Define colors
red = (239/255, 45/255, 81/255)
purple = (156/255, 39/255, 176/255)
orange = (255/255, 152/255, 0/255)
yellow = (255/255, 221/255, 0/255)

plt.hist(feature_importance_df['importance'], bins=50, edgecolor='black', color=red, alpha=0.7)
plt.xscale('symlog', linthresh=1e-5)  # Symmetric log scale
plt.xlabel('Feature Importance')
plt.ylabel('Frequency')
plt.title('Distribution of Feature Importance Scores')

mean_importance = feature_importance_df['importance'].mean()
median_importance = feature_importance_df['importance'].median()
top_30_threshold = feature_importance_df['importance'].nlargest(30).min()

plt.axvline(x=mean_importance, color=purple, linestyle='dashed', linewidth=2, label='Mean Importance')
plt.axvline(x=median_importance, color=orange, linestyle='dashed', linewidth=2, label='Median Importance')
plt.axvline(x=top_30_threshold, color=yellow, linestyle='dashed', linewidth=2, label='Top 30 Threshold')
plt.legend()
plt.grid(True, which="both", ls="-", alpha=0.2)
plt.tight_layout()
plt.savefig(os.path.join(output_dir, 'Part11_XGB_feature_importance_histogram.png'))
plt.close()
print("Finished task: Create histogram of feature importance scores")

print('XGBoost with feature importance calculation completed')